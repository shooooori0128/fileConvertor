#!/usr/bin/env python3
"""
Excel to Markdown Converter

Excelファイル(.xlsx)をMarkdown形式に変換するCLIアプリケーション
"""

import os
import sys
from pathlib import Path
import pandas as pd
import click
from openpyxl import load_workbook


class ExcelToMarkdownConverter:
    """ExcelからMarkdownへの変換を行うクラス"""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
    
    def validate_file(self, file_path: str) -> bool:
        """ファイルの存在と形式を確認"""
        path = Path(file_path)
        
        if not path.exists():
            click.echo(f"エラー: ファイル '{file_path}' が見つかりません。", err=True)
            return False
        
        if path.suffix.lower() not in self.supported_extensions:
            click.echo(f"エラー: サポートされていないファイル形式です。{self.supported_extensions}", err=True)
            return False
        
        return True
    
    def get_sheet_names(self, file_path: str) -> list:
        """Excelファイルのシート名一覧を取得"""
        try:
            workbook = load_workbook(file_path, read_only=True)
            return workbook.sheetnames
        except Exception as e:
            click.echo(f"エラー: シート名の取得に失敗しました - {e}", err=True)
            return []
    
    def convert_sheet_to_markdown(self, file_path: str, sheet_name: str) -> str:
        """指定されたシートをMarkdownテーブルに変換"""
        try:
            # Excelファイルを読み込み
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            
            # 空の行と列を削除
            df = df.dropna(how='all').dropna(axis=1, how='all')
            
            if df.empty:
                return f"\n## {sheet_name}\n\n*このシートにはデータがありません。*\n\n"
            
            # NaN値を空文字列に変換
            df = df.fillna('')
            
            # Markdownテーブル形式に変換
            markdown_table = df.to_markdown(index=False, tablefmt='github')
            
            return f"\n## {sheet_name}\n\n{markdown_table}\n\n"
            
        except Exception as e:
            click.echo(f"エラー: シート '{sheet_name}' の変換に失敗しました - {e}", err=True)
            return f"\n## {sheet_name}\n\n*変換エラー: {e}*\n\n"
    
    def convert_to_markdown(self, input_file: str, output_file: str = None, 
                          specific_sheet: str = None) -> bool:
        """ExcelファイルをMarkdownに変換"""
        
        if not self.validate_file(input_file):
            return False
        
        # 出力ディレクトリを確保
        output_dir = Path('./output')
        output_dir.mkdir(exist_ok=True)
        
        # 出力ファイル名を決定
        if output_file is None:
            input_path = Path(input_file)
            output_file = output_dir / input_path.with_suffix('.md').name
        else:
            # 出力ファイルが相対パスの場合、outputディレクトリに配置
            output_path = Path(output_file)
            if not output_path.is_absolute():
                output_file = output_dir / output_path
        
        try:
            # シート名を取得
            sheet_names = self.get_sheet_names(input_file)
            if not sheet_names:
                return False
            
            # 特定のシートが指定されている場合
            if specific_sheet:
                if specific_sheet not in sheet_names:
                    click.echo(f"エラー: シート '{specific_sheet}' が見つかりません。", err=True)
                    click.echo(f"利用可能なシート: {', '.join(sheet_names)}")
                    return False
                sheet_names = [specific_sheet]
            
            # Markdownコンテンツを生成
            markdown_content = f"# {Path(input_file).stem}\n\n"
            markdown_content += f"*Excel ファイル: {input_file}*\n"
            markdown_content += f"*変換日時: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}*\n"
            
            for sheet_name in sheet_names:
                click.echo(f"変換中: {sheet_name}")
                markdown_content += self.convert_sheet_to_markdown(input_file, sheet_name)
            
            # ファイルに書き出し
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            
            click.echo(f"変換完了: {output_file}")
            return True
            
        except Exception as e:
            click.echo(f"エラー: 変換処理に失敗しました - {e}", err=True)
            return False


@click.command()
@click.argument('input_file', type=click.Path(exists=True))
@click.option('-o', '--output', help='出力ファイル名（デフォルト: output/入力ファイル名.md）')
@click.option('-s', '--sheet', help='変換する特定のシート名')
@click.option('--list-sheets', is_flag=True, help='シート名一覧を表示して終了')
@click.version_option(version='1.0.0')
def main(input_file, output, sheet, list_sheets):
    """
    ExcelファイルをMarkdown形式に変換します。
    
    INPUT_FILE: 変換するExcelファイルのパス（推奨：input/ディレクトリ内）
    出力ファイルは自動的にoutput/ディレクトリに保存されます
    """
    converter = ExcelToMarkdownConverter()
    
    # シート一覧表示のみの場合
    if list_sheets:
        sheet_names = converter.get_sheet_names(input_file)
        if sheet_names:
            click.echo("利用可能なシート:")
            for name in sheet_names:
                click.echo(f"  - {name}")
        sys.exit(0)
    
    # 変換実行
    success = converter.convert_to_markdown(input_file, output, sheet)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
